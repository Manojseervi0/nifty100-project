from pathlib import Path
import sqlite3

import pandas as pd

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path("output")
OUTPUT_PATH.mkdir(exist_ok=True)

DB_PATH = "nifty100.db"


def load_peer_report_data():
    """
    Load peer comparison report data from SQLite.
    """

    conn = sqlite3.connect(DB_PATH)

    query = """
    SELECT

        pg.peer_group_name,
        pg.company_id,
        pg.is_benchmark,

        fr.year,
        fr.return_on_equity_pct,
        fr.return_on_capital_employed_pct,
        fr.net_profit_margin_pct,
        fr.operating_profit_margin_pct,
        fr.debt_to_equity,
        fr.interest_coverage,
        fr.asset_turnover,
        fr.free_cash_flow_cr,
        fr.revenue_cagr_5yr,
        fr.pat_cagr_5yr,
        fr.eps_cagr_5yr,
        fr.composite_quality_score,

        pp.metric,
        pp.percentile_rank

    FROM peer_groups pg

    LEFT JOIN financial_ratios fr
        ON pg.company_id = fr.company_id

    LEFT JOIN peer_percentiles pp
        ON pg.company_id = pp.company_id
       AND fr.year = pp.year
    """

    df = pd.read_sql(query, conn)

    conn.close()

    df = (
        df.sort_values("year")
          .drop_duplicates(
              subset=["company_id", "metric"],
              keep="last"
          )
    )

    return df


def prepare_peer_report(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare one row per company with KPI and percentile columns.
    """

    kpi_df = (
        df.drop(columns=["metric", "percentile_rank"])
          .drop_duplicates(subset=["company_id"])
          .reset_index(drop=True)
    )

    percentile_df = (
        df.pivot(
            index="company_id",
            columns="metric",
            values="percentile_rank"
        )
        .add_suffix("_percentile")
        .reset_index()
    )

    report = kpi_df.merge(
        percentile_df,
        on="company_id",
        how="left"
    )

    report = report.drop(
        columns=["nan_percentile"],
        errors="ignore"
    )

    return report


def export_peer_report():
    """
    Export peer comparison report to Excel.
    """

    report = prepare_peer_report(
        load_peer_report_data()
    )

    output_file = OUTPUT_PATH / "peer_comparison.xlsx"

    green = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )

    yellow = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C"
    )

    red = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )

    gold = PatternFill(
        fill_type="solid",
        fgColor="FFD966"
    )

    percentile_columns = [
        c
        for c in report.columns
        if c.endswith("_percentile")
    ]

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        peer_groups = sorted(
            report["peer_group_name"]
            .dropna()
            .unique()
        )

        for group in peer_groups:

            sheet_df = report[
                report["peer_group_name"] == group
            ].copy()

            sheet_df.to_excel(
                writer,
                sheet_name=group[:31],
                index=False
            )

            ws = writer.sheets[group[:31]]

            # ------------------------
            # Header
            # ------------------------

            for cell in ws[1]:
                cell.font = Font(bold=True)

            ws.freeze_panes = "A2"

            # ------------------------
            # Auto width
            # ------------------------

            for column in ws.columns:

                width = max(
                    len(str(cell.value))
                    if cell.value is not None else 0
                    for cell in column
                )

                ws.column_dimensions[
                    get_column_letter(column[0].column)
                ].width = min(width + 2, 30)

            # ------------------------
            # Column lookup
            # ------------------------

            column_map = {
                cell.value: cell.column
                for cell in ws[1]
            }

            # ------------------------
            # Benchmark highlight
            # ------------------------

            if "is_benchmark" in column_map:

                benchmark_col = column_map["is_benchmark"]

                for row in range(2, ws.max_row + 1):

                    if ws.cell(row, benchmark_col).value == 1:

                        for cell in ws[row]:
                            cell.fill = gold

            # ------------------------
            # Percentile colouring
            # ------------------------

            for column_name in percentile_columns:

                if column_name not in column_map:
                    continue

                col = column_map[column_name]

                for row in range(2, ws.max_row + 1):

                    value = ws.cell(row, col).value

                    if value is None:
                        continue

                    try:
                        value = float(value)
                    except (TypeError, ValueError):
                        continue

                    if value >= 0.75:
                        ws.cell(row, col).fill = green

                    elif value <= 0.25:
                        ws.cell(row, col).fill = red

                    else:
                        ws.cell(row, col).fill = yellow

            # ------------------------
            # Median Row
            # ------------------------

            median_row = ["Median"]

            for column in sheet_df.columns[1:]:

                if pd.api.types.is_numeric_dtype(
                    sheet_df[column]
                ):

                    median_row.append(
                        sheet_df[column].median()
                    )

                else:

                    median_row.append("")

            ws.append(median_row)

    print(f"Exported {output_file}")


if __name__ == "__main__":
    export_peer_report()